import os
import sys
import re
import shutil
from os import path
from os import walk
from openpyxl import load_workbook


def check_arguments():
    if len(sys.argv) != 3:
        print("Usage:")
        print("python a_celex_extractor.py <folder_with_xlsx_files> <output_directory>")
        exit(1)


def append_to_file(target_directory: str, celex: str):
    """
    Appending celex to some file inside the target directory
    :param target_directory:
    :param celex:
    :return:
    """
    celex_part = celex[0:6]

    directory_path = path.join(target_directory, celex_part)
    if not path.exists(directory_path):
        os.mkdir(directory_path)

    file_path = path.join(target_directory, celex_part, f"{celex_part}.csv")
    # check if a given record exists first
    record_exists = False
    if path.exists(file_path):
        with open(file_path, "r", encoding="utf-8") as src:
            for line in src.readlines():
                if line is None:
                    continue
                line = line.replace("\n", "")
                if line == celex:
                    record_exists = True
    if not record_exists:
        with open(file_path, "a", encoding="utf-8") as target:
            target.write(celex)
            target.write("\n")


def read_file(excel_file: str, target_directory: str):
    """
    Reading a single XLS(X) file and extracting celex from all the tabs
    :param excel_file:
    :param target_directory:
    :return:
    """
    with_celex_pattern = re.compile("Celex No\.\s(.*)")
    celex_pattern = re.compile("^(1\d{4}[a-zA-Z].*)")
    print(f"Reading file {excel_file}")
    wb = load_workbook(excel_file)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Reading content of the sheet {sheet_name}")
        empty_count = 0
        for row in sheet.rows:
            all_empty = True
            for cell in row:
                if cell.value is not None:
                    all_empty = False
                    if not with_celex_pattern.search(cell.value) is None:
                        celex_numbers = with_celex_pattern.findall(cell.value)
                        for celex in celex_numbers:
                            append_to_file(target_directory, celex)
                    elif not celex_pattern.search(cell.value) is None:
                        celex_numbers = celex_pattern.findall(cell.value)
                        for celex in celex_numbers:
                            append_to_file(target_directory, celex)
            if all_empty:
                empty_count += 1
            if empty_count > 5:
                break


def read_folder(folder_path: str, target_directory: str):
    """
    Reading all the excel files from the given folder, extracting
    celex numbers from files

    :param folder_path:
    :param target_directory:
    :return:
    """
    if not path.exists(target_directory):
        os.mkdir(target_directory)

    for (dir_path, dir_names, file_names) in walk(folder_path):
        for source_file in file_names:
            if ".xls" not in source_file:
                print(f"{source_file} is not an excel doc, skipping")
                continue

            excel_file = path.join(folder_path, source_file)
            read_file(excel_file, target_directory)


if __name__ == '__main__':
    check_arguments()
    read_folder(sys.argv[1], sys.argv[2])
